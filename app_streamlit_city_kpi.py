#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import csv
import io
import posixpath
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Tuple

import streamlit as st

try:
    import pandas as pd  # Optional; used if available for nicer tables.
except Exception:  # pragma: no cover
    pd = None

EXCEL_NS_MAIN = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
EXCEL_NS_REL = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}

SUMMARY_COLUMNS = [
    "城市",
    "期间",
    "产成率(%)",
    "产值(元/kg)",
    "生肉产量(吨)",
    "销量(吨)",
    "产销率(%)",
    "宰鸡量(千只)",
    "均重(kg/只)",
]

PART_ORDER = [
    "腿类",
    "胸部",
    "大胸",
    "胸皮",
    "里肌",
    "翅类",
    "骨架",
    "爪类",
    "鸡肝",
    "鸡心",
    "脖类",
    "鸡胗",
    "下料类",
    "鸡头",
    "油类",
    "骨架+鸡脖+鸡头",
    "骨架+鸡脖",
    "主产品",
    "副产品",
    "总计",
]

CITY_ORDER = ["大连", "铁岭", "沧州", "蚌埠"]
AUTO_LINE_CITIES = {"大连", "蚌埠"}
NONAUTO_LINE_CITIES = {"沧州", "铁岭"}

STANDARD_RATE_COLUMNS = ["自动掏膛线工厂标准", "非自动掏膛线工厂标准"]
STANDARD_RATE_VALUES = [
    (0.2550, 0.2540),
    (0.2170, 0.2140),
    (0.1940, 0.1910),
    (0.0230, 0.0230),
    (0.0370, 0.0370),
    (0.0850, 0.0840),
    (0.1500, 0.1550),
    (0.0350, 0.0350),
    (0.0190, 0.0200),
    (0.0043, 0.0045),
    (0.0550, 0.0600),
    (0.0100, 0.0100),
    (0.0200, 0.0100),
    (0.0197, 0.0200),
    (0.0230, 0.0215),
    (0.2247, 0.2350),
    (0.2050, 0.2150),
    (0.6290, 0.6240),
    (0.3010, 0.3010),
    (0.9300, 0.9250),
]
PART_STANDARD_RATES = {part: vals for part, vals in zip(PART_ORDER, STANDARD_RATE_VALUES)}

PART_ALIASES = {
    "胸部": ["胸类", "胸部"],
    "大胸": ["胸类-胸", "大胸"],
    "胸皮": ["胸类-胸皮", "胸皮"],
    "里肌": ["里肌类", "里肌"],
    "骨架": ["骨架类", "骨架"],
    "爪类": ["爪类"],
    "鸡肝": ["鸡肝类", "鸡肝"],
    "鸡心": ["鸡心类", "鸡心"],
    "鸡胗": ["鸡胗类", "鸡胗"],
    "鸡头": ["鸡头类", "鸡头"],
    "脖类": ["脖类", "鸡脖"],
    "油类": ["油类"],
    "下料类": ["下料类", "其他内脏"],
    "骨架+鸡脖+鸡头": ["鸡头+鸡脖+骨架", "骨架+鸡脖+鸡头"],
    "骨架+鸡脖": ["骨架+鸡脖"],
    "腿类": ["腿类"],
}

DERIVED_PARTS = {
    "骨架+鸡脖+鸡头": ["骨架", "脖类", "鸡头"],
    "骨架+鸡脖": ["骨架", "脖类"],
    "主产品": ["腿类", "胸类", "里肌", "翅类", "爪类"],
}

HEADER_ALIASES = {
    "日期": "期间",
    "产成率": "产成率(%)",
    "产成率(%)": "产成率(%)",
    "产值": "产值(元/kg)",
    "产值(元/kg)": "产值(元/kg)",
    "生肉产量": "生肉产量(吨)",
    "生肉产量(吨)": "生肉产量(吨)",
    "产量": "生肉产量(吨)",
    "销量": "销量(吨)",
    "销量(吨)": "销量(吨)",
    "产销率": "产销率(%)",
    "产销率(%)": "产销率(%)",
    "宰鸡量": "宰鸡量(千只)",
    "宰鸡量(千只)": "宰鸡量(千只)",
    "均重": "均重(kg/只)",
    "均重(kg/只)": "均重(kg/只)",
}

CHART_PATH = Path("/home/fan/8f573a0c22d5ad3933140a75507eacb8.png")


def read_shared_strings(zip_file: zipfile.ZipFile) -> List[str]:
    if "xl/sharedStrings.xml" not in zip_file.namelist():
        return []
    root = ET.fromstring(zip_file.read("xl/sharedStrings.xml"))
    strings: List[str] = []
    for si in root.findall("m:si", EXCEL_NS_MAIN):
        text = "".join(t.text or "" for t in si.findall(".//m:t", EXCEL_NS_MAIN))
        strings.append(text)
    return strings


def column_to_index(col: str) -> int:
    idx = 0
    for ch in col:
        idx = idx * 26 + (ord(ch.upper()) - ord("A") + 1)
    return idx - 1


def read_first_sheet_bytes(data: bytes) -> Tuple[List[str], List[List[str]]]:
    """Return header row + dense rows from the first worksheet in bytes."""
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        shared = read_shared_strings(zf)
        workbook = ET.fromstring(zf.read("xl/workbook.xml"))
        rels_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rels = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels_root.findall("rel:Relationship", EXCEL_NS_REL)}
        first_sheet = workbook.find("m:sheets/m:sheet", EXCEL_NS_MAIN)
        if first_sheet is None:
            return [], []
        rel_id = first_sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
        target = rels.get(rel_id, "worksheets/sheet1.xml")
        sheet_path = posixpath.normpath(posixpath.join("xl", target))
        root = ET.fromstring(zf.read(sheet_path))

        rows: List[Dict[int, str]] = []
        max_col = -1
        for row in root.findall("m:sheetData/m:row", EXCEL_NS_MAIN):
            values: Dict[int, str] = {}
            for cell in row.findall("m:c", EXCEL_NS_MAIN):
                ref = cell.attrib.get("r", "")
                col_letters = "".join(ch for ch in ref if ch.isalpha())
                if not col_letters:
                    continue
                col_idx = column_to_index(col_letters)
                max_col = max(max_col, col_idx)
                cell_type = cell.attrib.get("t")
                value = ""
                v = cell.find("m:v", EXCEL_NS_MAIN)
                inline = cell.find("m:is", EXCEL_NS_MAIN)
                if cell_type == "s" and v is not None and v.text is not None:
                    s_idx = int(v.text)
                    if 0 <= s_idx < len(shared):
                        value = shared[s_idx]
                elif cell_type == "inlineStr" and inline is not None:
                    value = "".join(t.text or "" for t in inline.findall(".//m:t", EXCEL_NS_MAIN))
                elif v is not None and v.text is not None:
                    value = v.text
                if value != "":
                    values[col_idx] = value
            rows.append(values)
        width = max_col + 1 if max_col >= 0 else 0
        dense_rows: List[List[str]] = []
        for row in rows:
            dense_rows.append([row.get(i, "") for i in range(width)])
        header = dense_rows[0] if dense_rows else []
        return header, dense_rows


def _find_titled_table_header(
    rows: List[List[str]],
    title: str,
    header_label: str = "项目",
) -> Tuple[int, int] | None:
    for row_idx, row in enumerate(rows):
        for col_idx, cell in enumerate(row):
            if str(cell).strip() != title:
                continue
            for j in range(row_idx + 1, min(row_idx + 6, len(rows))):
                if col_idx < len(rows[j]) and str(rows[j][col_idx]).strip() == header_label:
                    return j, col_idx
    return None


def normalize_date(text: str) -> str:
    cleaned = (text or "").strip()
    if not cleaned:
        return ""
    try:
        as_float = float(cleaned.replace(",", ""))
        if as_float.is_integer():
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=int(as_float))).date().isoformat()
    except ValueError:
        pass
    return cleaned


def build_column_map(header_row: List[str]) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for idx, name in enumerate(header_row):
        normalized = HEADER_ALIASES.get((name or "").strip())
        if normalized:
            mapping[normalized] = idx
    return mapping


def extract_records_from_bytes(data: bytes, filename: str) -> List[Dict[str, str]]:
    _, rows = read_first_sheet_bytes(data)
    header_idx = None
    for idx, row in enumerate(rows):
        if any(cell.strip() == "日期" for cell in row):
            header_idx = idx
            break
    if header_idx is None:
        return []

    header_row = rows[header_idx]
    col_map = build_column_map(header_row)
    data_rows = rows[header_idx + 1 :]
    records: List[Dict[str, str]] = []
    city = Path(filename).stem.split("_", 1)[0]

    for row in data_rows:
        if not any(cell.strip() for cell in row):
            continue
        period_idx = col_map.get("期间")
        period = normalize_date(row[period_idx]) if period_idx is not None and period_idx < len(row) else ""
        values: Dict[str, str] = {"城市": city, "期间": period}
        empty_value = True
        for col in SUMMARY_COLUMNS[2:]:
            idx = col_map.get(col)
            value = row[idx].strip() if idx is not None and idx < len(row) else ""
            values[col] = value
            if value:
                empty_value = False
        if period or not empty_value:
            records.append(values)
    return records


def to_csv_bytes(records: List[Dict[str, str]]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=SUMMARY_COLUMNS)
    writer.writeheader()
    writer.writerows(records)
    return buf.getvalue().encode("utf-8")


def to_csv_bytes_generic(records: List[Dict[str, str]]) -> bytes:
    if not records:
        return "".encode("utf-8")
    fields = sorted({k for r in records for k in r.keys()})
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fields)
    writer.writeheader()
    writer.writerows(records)
    return buf.getvalue().encode("utf-8")


def extract_monthly_value_from_bytes(data: bytes, filename: str) -> List[Dict[str, str]]:
    """优先读取“组合还原后产成率总览（月累计）”，缺失则回退“本月至今累计/月累计”，取含税单价/金额/产量。"""
    _, rows = read_first_sheet_bytes(data)

    def _read_block(marker: str) -> List[Dict[str, str]]:
        start_idx = None
        for idx, row in enumerate(rows):
            if row and str(row[0]).strip() == marker:
                start_idx = idx
                break
        if start_idx is None:
            return []
        header_idx = None
        for j in range(start_idx + 1, min(start_idx + 5, len(rows))):
            if rows[j] and str(rows[j][0]).strip() == "项目":
                header_idx = j
                break
        if header_idx is None:
            return []

        header = rows[header_idx]
        name_idx = 0
        volume_idx = _find_col_idx(header, ["产量(kg)", "产量", "生产量", "产量kg", "产量KG", "产出量"])
        if volume_idx is None:
            volume_idx = _find_col_idx(header, ["销量(kg)", "销量", "销售量", "销量kg", "销量KG"])
        amount_idx = _find_col_idx(header, ["含税金额", "含税金额(元)", "金额"])
        unit_idx = _find_col_idx(header, ["含税单价", "单价"])
        if unit_idx is None and (amount_idx is None or volume_idx is None):
            return []

        stop_words = {"部位还原后的产成率", "组合还原后产成率总览（月累计）", "组合还原后产成率总览"}
        data_rows = rows[header_idx + 1 :]
        plant = Path(filename).stem.split("_", 1)[0]
        out: List[Dict[str, str]] = []
        for row in data_rows:
            if not any(str(c).strip() for c in row):
                break
            first = str(row[name_idx]).strip() if len(row) > name_idx else ""
            if first in stop_words:
                break
            if not first:
                continue
            part = normalize_part_name(first)
            volume = _parse_number(row[volume_idx]) if (volume_idx is not None and volume_idx < len(row)) else None
            amount = _parse_number(row[amount_idx]) if (amount_idx is not None and amount_idx < len(row)) else None
            unit = _parse_number(row[unit_idx]) if (unit_idx is not None and unit_idx < len(row)) else None
            if unit is None and amount is not None and volume is not None and volume > 0:
                unit = amount / volume
            if unit is None and amount is None and volume is None:
                continue
            out.append({"城市": plant, "部位": part, "产量": volume, "含税金额": amount, "含税单价": unit})
        return out

    return (
        _read_block("组合还原后产成率总览（月累计）")
        or _read_block("本月至今累计")
        or _read_month_total_side_table(rows, filename)
    )


def _read_month_total_side_table(rows: List[List[str]], filename: str) -> List[Dict[str, str]]:
    loc = _find_titled_table_header(rows, "月累计")
    if not loc:
        return []
    header_idx, start_col = loc
    header = rows[header_idx][start_col:]
    name_idx = 0
    volume_idx = _find_col_idx(header, ["产量(kg)", "产量", "生产量", "产量kg", "产量KG", "产出量"])
    if volume_idx is None:
        volume_idx = _find_col_idx(header, ["销量(kg)", "销量", "销售量", "销量kg", "销量KG"])
    amount_idx = _find_col_idx(header, ["含税金额", "含税金额(元)", "金额"])
    unit_idx = _find_col_idx(header, ["含税单价", "单价"])
    if unit_idx is None and (amount_idx is None or volume_idx is None):
        return []
    last_idx = max(idx for idx in [name_idx, volume_idx, amount_idx, unit_idx] if idx is not None)
    plant = Path(filename).stem.split("_", 1)[0]
    out: List[Dict[str, str]] = []
    for row in rows[header_idx + 1 :]:
        row_slice = row[start_col : start_col + last_idx + 1]
        if not any(str(c).strip() for c in row_slice):
            break
        first = str(row[start_col + name_idx]).strip() if start_col + name_idx < len(row) else ""
        if not first:
            continue
        part = normalize_part_name(first)
        volume = _parse_number(row[start_col + volume_idx]) if (volume_idx is not None and start_col + volume_idx < len(row)) else None
        amount = _parse_number(row[start_col + amount_idx]) if (amount_idx is not None and start_col + amount_idx < len(row)) else None
        unit = _parse_number(row[start_col + unit_idx]) if (unit_idx is not None and start_col + unit_idx < len(row)) else None
        if unit is None and amount is not None and volume is not None and volume > 0:
            unit = amount / volume
        if unit is None and amount is None and volume is None:
            continue
        out.append({"城市": plant, "部位": part, "产量": volume, "含税金额": amount, "含税单价": unit})
    return out


def _read_month_total_parts_side_table(rows: List[List[str]], filename: str) -> List[Dict[str, str]]:
    loc = _find_titled_table_header(rows, "月累计")
    if not loc:
        return []
    header_idx, start_col = loc
    header = rows[header_idx][start_col:]
    rate_idx = _find_col_idx(
        header,
        ["调整后产成率(%)", "调整后产成率", "产成率(%)", "产成率%", "产成率"],
    )
    if rate_idx is None:
        return []
    last_idx = max(0, rate_idx)
    plant = Path(filename).stem.split("_", 1)[0]
    out: List[Dict[str, str]] = []
    for row in rows[header_idx + 1 :]:
        row_slice = row[start_col : start_col + last_idx + 1]
        if not any(str(c).strip() for c in row_slice):
            break
        first = str(row[start_col]).strip() if start_col < len(row) else ""
        if not first:
            continue
        rate_val = row[start_col + rate_idx] if start_col + rate_idx < len(row) else ""
        out.append({"城市": plant, "部位": normalize_part_name(first), "产成率": rate_val})
    return out


def _find_col_idx(header: List[str], name_cands: List[str]) -> int | None:
    clean = [str(c or "").strip() for c in header]
    for cand in name_cands:
        if cand in clean:
            return clean.index(cand)
    # loose contains
    for i, col in enumerate(clean):
        for cand in name_cands:
            if cand and cand in col:
                return i
    return None


def extract_monthly_parts_from_bytes(data: bytes, filename: str) -> List[Dict[str, str]]:
    """优先读取“组合还原后产成率总览（月累计）”，若缺失回退“本月至今累计/月累计”."""
    header_row, rows = read_first_sheet_bytes(data)

    def _read_block(marker: str) -> List[Dict[str, str]]:
        start_idx = None
        for idx, row in enumerate(rows):
            if row and str(row[0]).strip() == marker:
                start_idx = idx
                break
        if start_idx is None:
            return []
        header_idx = None
        for j in range(start_idx + 1, min(start_idx + 5, len(rows))):
            if rows[j] and str(rows[j][0]).strip() == "项目":
                header_idx = j
                break
        if header_idx is None:
            return []
        header = rows[header_idx]
        rate_idx = _find_col_idx(
            header,
            ["调整后产成率(%)", "调整后产成率", "产成率(%)", "产成率%", "产成率"],
        )
        name_idx = 0
        stop_words = {"组合还原后产成率总览（月累计）", "组合还原后产成率总览"}
        data_rows = rows[header_idx + 1 :]
        out: List[Dict[str, str]] = []
        plant = Path(filename).stem.split("_", 1)[0]
        for row in data_rows:
            if not any(str(c).strip() for c in row):
                break
            first = str(row[name_idx]).strip() if len(row) > name_idx else ""
            if first in stop_words:
                break
            if not first:
                continue
            rate_val = row[rate_idx] if (rate_idx is not None and rate_idx < len(row)) else ""
            out.append({"城市": plant, "部位": normalize_part_name(first), "产成率": rate_val})
        return out

    return (
        _read_block("组合还原后产成率总览（月累计）")
        or _read_block("本月至今累计")
        or _read_month_total_parts_side_table(rows, filename)
    )


def _parse_number(val: str) -> float | None:
    text = str(val or "").replace(",", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _parse_percent(val: str) -> float | None:
    text = str(val or "").strip()
    if not text:
        return None
    has_percent = "%" in text
    text = text.replace("%", "").replace(",", "").strip()
    if not text:
        return None
    try:
        num = float(text)
    except ValueError:
        return None
    if has_percent:
        return num / 100
    if abs(num) > 2:
        return num / 100
    return num


def normalize_part_name(name: str) -> str:
    cleaned = str(name or "").strip()
    for target, aliases in PART_ALIASES.items():
        for a in aliases:
            if cleaned == a:
                return target
    return cleaned


def _normalize_parts_list(parts: List[str]) -> List[str]:
    return [normalize_part_name(p) for p in parts]


def _build_rate_lookup(part_records: List[Dict[str, str]]) -> Dict[str, Dict[str, float]]:
    rate_lookup: Dict[str, Dict[str, float]] = {}
    for record in part_records:
        part = normalize_part_name(record.get("部位"))
        city = record.get("城市")
        rate = _parse_percent(record.get("产成率"))
        if part and city and rate is not None:
            rate_lookup.setdefault(city, {}).setdefault(part, 0.0)
            rate_lookup[city][part] += rate
    return rate_lookup


def _format_thousands0(val: str) -> str:
    num = _parse_number(val)
    return f"{num:,.0f}" if num is not None else ""


def _format_percent2(val: str) -> str:
    num = _parse_percent(val)
    return f"{num * 100:.2f}%" if num is not None else ""


def _format_percent0(val: str) -> str:
    num = _parse_percent(val)
    return f"{num * 100:.0f}%" if num is not None else ""


def _format_decimal2(val: float | str | None) -> str:
    num = _parse_number(val)
    return f"{num:.2f}" if num is not None else ""


SUMMARY_EXPORT_FIELDS = [
    ("工厂", "城市", None),
    ("日期", "期间", _parse_number),
    ("产成率(%)", "产成率(%)", _parse_percent),
    ("产值(元/kg)", "产值(元/kg)", _parse_number),
    ("生肉产量(吨)", "生肉产量(吨)", _parse_number),
    ("销量(吨)", "销量(吨)", _parse_number),
    ("产销率(%)", "产销率(%)", _parse_percent),
    ("宰鸡量(千只)", "宰鸡量(千只)", _parse_number),
    ("均重(kg/只)", "均重(kg/只)", _parse_number),
]

REPORT_NOTE_RIGHT = "备注：主产品、副产品为单位产品所含该项价值。"
REPORT_NOTES_LEFT = [
    "备注：1、主产品包括（腿、胸、翅、里肌、爪）",
    "          2、低于标准的数值为绿色",
    "          3、产成为整鸡还原之后的产成",
    "          4、红色数字为最大值，蓝色数字为最小值",
    "          5、胸部为大胸与胸皮的合计值",
]


def _coerce_value(val: str, parser) -> float | str:
    if parser is None:
        return str(val or "").strip()
    num = parser(val)
    if num is not None:
        return num
    return str(val or "").strip()


def _build_part_pivot_map(records: List[Dict[str, str]]) -> Tuple[Dict[str, Dict[str, float]], List[str]]:
    if not records:
        return {}, []
    if pd:
        dfp = pd.DataFrame(records)
        dfp["产成率"] = dfp["产成率"].apply(_parse_percent)
        dfp = dfp.dropna(subset=["产成率"])
        dfp["部位"] = dfp["部位"].apply(normalize_part_name)
        dfp = dfp.groupby(["部位", "城市"], as_index=False)["产成率"].sum()
        pivot = dfp.pivot_table(index="部位", columns="城市", values="产成率", aggfunc="sum")
        for new_part, parts in DERIVED_PARTS.items():
            parts_norm = _normalize_parts_list(parts)
            vals = {}
            for city in pivot.columns:
                vals[city] = sum(
                    pivot.at[p, city] if p in pivot.index and pd.notna(pivot.at[p, city]) else 0 for p in parts_norm
                )
            pivot.loc[new_part] = vals
        if "总计" not in pivot.index:
            pivot.loc["总计"] = pivot.sum(min_count=1)
        if "总计" in pivot.index and "主产品" in pivot.index:
            pivot.loc["副产品"] = pivot.loc["总计"] - pivot.loc["主产品"]
        pivot = pivot.reindex(PART_ORDER, fill_value=pd.NA)
        part_map = {}
        for part, row in pivot.iterrows():
            part_map[part] = {city: (None if pd.isna(val) else float(val)) for city, val in row.items()}
        return part_map, _order_cities(list(pivot.columns))

    parts = {}
    for r in records:
        part = normalize_part_name(r.get("部位"))
        city = r.get("城市")
        val = _parse_percent(r.get("产成率"))
        if part and city and val is not None:
            parts.setdefault(part, {}).setdefault(city, 0.0)
            parts[part][city] += val

    all_cities = {c for v in parts.values() for c in v}
    for new_part, comp in DERIVED_PARTS.items():
        comp_norm = _normalize_parts_list(comp)
        city_vals = {}
        for city in all_cities:
            city_vals[city] = sum(parts.get(p, {}).get(city, 0) for p in comp_norm)
        parts[new_part] = city_vals

    if "总计" not in parts:
        parts["总计"] = {c: sum(parts.get(p, {}).get(c, 0) for p in parts if p != "总计") for c in all_cities}
    if "总计" in parts and "主产品" in parts:
        sub_city = {}
        for city in all_cities:
            total_val = parts.get("总计", {}).get(city)
            main_val = parts.get("主产品", {}).get(city)
            if total_val is None or main_val is None:
                continue
            sub_city[city] = total_val - main_val
        if sub_city:
            parts["副产品"] = sub_city

    return parts, _order_cities(list(all_cities))


def _build_value_pivot_map(
    records: List[Dict[str, str]],
    part_records: List[Dict[str, str]] | None = None,
) -> Tuple[Dict[str, Dict[str, float]], List[str]]:
    if not records:
        return {}, []
    rate_lookup = _build_rate_lookup(part_records or [])
    if pd:
        dfv = pd.DataFrame(records)
        dfv["含税单价"] = dfv.apply(
            lambda r: r["含税单价"]
            if pd.notna(r.get("含税单价"))
            else (r["含税金额"] / r["产量"] if pd.notna(r.get("含税金额")) and pd.notna(r.get("产量")) and r["产量"] > 0 else pd.NA),
            axis=1,
        )
        dfv = dfv.dropna(subset=["含税单价"])
        dfv["部位"] = dfv["部位"].apply(normalize_part_name)
        base = dfv[["部位", "城市", "含税单价", "产量", "含税金额"]]
        def _agg_base(group):
            volume = group["产量"].sum(min_count=1)
            amount = group["含税金额"].sum(min_count=1)
            weighted_amount = (group["含税单价"] * group["产量"]).sum(min_count=1)
            unit = pd.NA
            if pd.notna(volume) and volume > 0:
                if pd.notna(amount):
                    unit = amount / volume
                elif pd.notna(weighted_amount):
                    unit = weighted_amount / volume
            elif group["含税单价"].notna().any():
                unit = group["含税单价"].mean()
            return pd.Series({"含税单价": unit, "产量": volume, "含税金额": amount})

        base = base.groupby(["部位", "城市"]).apply(_agg_base).reset_index()

        derived_rows = []
        for new_part, parts in DERIVED_PARTS.items():
            parts_norm = _normalize_parts_list(parts)
            if new_part == "主产品" and rate_lookup:
                for city, part_rates in rate_lookup.items():
                    numerator = 0.0
                    denominator = 0.0
                    rate_sum = 0.0
                    total_rate = part_rates.get("总计")
                    for part in parts_norm:
                        rate = part_rates.get(part)
                        if rate is None:
                            continue
                        unit_series = base.loc[(base["部位"] == part) & (base["城市"] == city), "含税单价"]
                        if unit_series.empty:
                            continue
                        unit_val = unit_series.iloc[0]
                        if pd.isna(unit_val):
                            continue
                        numerator += float(unit_val) * rate
                        rate_sum += rate
                    denominator = total_rate if total_rate not in (None, 0) else rate_sum
                    if denominator > 0:
                        derived_rows.append(
                            {
                                "部位": new_part,
                                "城市": city,
                                "含税单价": numerator / denominator,
                                "产量": pd.NA,
                                "含税金额": pd.NA,
                            }
                        )
                continue
            sub = base[base["部位"].isin(parts_norm)]
            if sub.empty:
                continue
            unit_sub = sub.dropna(subset=["含税单价", "产量"]).copy()
            if not unit_sub.empty:
                unit_sub["加权金额"] = unit_sub["含税单价"] * unit_sub["产量"]
                agg = unit_sub.groupby("城市").agg({"加权金额": "sum", "产量": "sum"})
                agg["含税单价"] = agg.apply(
                    lambda r: r["加权金额"] / r["产量"] if pd.notna(r["产量"]) and r["产量"] > 0 else pd.NA,
                    axis=1,
                )
                for city, row in agg.iterrows():
                    derived_rows.append(
                        {
                            "部位": new_part,
                            "城市": city,
                            "含税单价": row["含税单价"],
                            "产量": row["产量"],
                            "含税金额": row["加权金额"],
                        }
                    )
                continue
            amt_sub = sub.dropna(subset=["含税金额", "产量"])
            if amt_sub.empty:
                continue
            agg = amt_sub.groupby("城市").agg({"含税金额": "sum", "产量": "sum"})
            agg["含税单价"] = agg.apply(
                lambda r: r["含税金额"] / r["产量"] if pd.notna(r["产量"]) and r["产量"] > 0 else pd.NA,
                axis=1,
            )
            for city, row in agg.iterrows():
                derived_rows.append(
                    {
                        "部位": new_part,
                        "城市": city,
                        "含税单价": row["含税单价"],
                        "产量": row["产量"],
                        "含税金额": row["含税金额"],
                    }
                )

        if derived_rows:
            derived_df = pd.DataFrame(derived_rows)
            base = base.merge(derived_df[["部位", "城市"]], on=["部位", "城市"], how="left", indicator=True)
            base = base[base["_merge"] == "left_only"].drop(columns=["_merge"])
            combined = pd.concat([base, derived_df], ignore_index=True)
        else:
            combined = base

        if "总计" not in combined["部位"].values:
            total = base.groupby("城市").agg({"含税金额": "sum", "产量": "sum"})
            total["含税单价"] = total.apply(
                lambda r: r["含税金额"] / r["产量"] if pd.notna(r["产量"]) and r["产量"] > 0 else pd.NA,
                axis=1,
            )
            for city, row in total.iterrows():
                combined = pd.concat(
                    [
                        combined,
                        pd.DataFrame(
                            [
                                {
                                    "部位": "总计",
                                    "城市": city,
                                    "含税单价": row["含税单价"],
                                    "产量": row["产量"],
                                    "含税金额": row["含税金额"],
                                }
                            ]
                        ),
                    ],
                    ignore_index=True,
                )

        pivot = combined.pivot_table(index="部位", columns="城市", values="含税单价", aggfunc="first")
        if "总计" in pivot.index and "主产品" in pivot.index:
            pivot.loc["副产品"] = pivot.loc["总计"] - pivot.loc["主产品"]
        pivot = pivot.reindex(PART_ORDER, fill_value=pd.NA)
        value_map = {}
        for part, row in pivot.iterrows():
            value_map[part] = {city: (None if pd.isna(val) else float(val)) for city, val in row.items()}
        return value_map, _order_cities(list(pivot.columns))

    parts_map: Dict[str, Dict[str, Dict[str, float]]] = {}
    for r in records:
        part = normalize_part_name(r.get("部位"))
        city = r.get("城市")
        if not part or not city:
            continue
        volume = _parse_number(r.get("产量"))
        amount = _parse_number(r.get("含税金额"))
        unit = _parse_number(r.get("含税单价"))
        if unit is None and amount is not None and volume is not None and volume > 0:
            unit = amount / volume
        if unit is None:
            continue
        parts_map.setdefault(part, {}).setdefault(city, {"amount": 0.0, "volume": 0.0})
        if amount is not None:
            parts_map[part][city]["amount"] += amount
        if volume is not None:
            parts_map[part][city]["volume"] += volume
        parts_map[part][city]["unit"] = unit

    all_cities = {c for v in parts_map.values() for c in v}
    for new_part, comp in DERIVED_PARTS.items():
        comp_norm = _normalize_parts_list(comp)
        if new_part == "主产品" and rate_lookup:
            city_vals = {}
            for city in all_cities:
                numerator = 0.0
                denominator = 0.0
                rate_sum = 0.0
                total_rate = rate_lookup.get(city, {}).get("总计")
                amt = 0.0
                vol = 0.0
                for part in comp_norm:
                    part_data = parts_map.get(part, {}).get(city, {})
                    unit = part_data.get("unit")
                    rate = rate_lookup.get(city, {}).get(part)
                    if rate is not None and unit is not None:
                        numerator += unit * rate
                        rate_sum += rate
                    amt += part_data.get("amount", 0)
                    vol += part_data.get("volume", 0)
                denominator = total_rate if total_rate not in (None, 0) else rate_sum
                if denominator > 0:
                    city_vals[city] = {"amount": amt, "volume": vol, "unit": numerator / denominator}
            if city_vals:
                parts_map[new_part] = city_vals
            continue
        city_vals = {}
        for city in all_cities:
            amt = sum(parts_map.get(p, {}).get(city, {}).get("amount", 0) for p in comp_norm)
            vol = sum(parts_map.get(p, {}).get(city, {}).get("volume", 0) for p in comp_norm)
            unit = amt / vol if vol > 0 else None
            if unit is not None:
                city_vals[city] = {"amount": amt, "volume": vol, "unit": unit}
        if city_vals:
            parts_map[new_part] = city_vals

    if "总计" not in parts_map:
        total_city = {}
        for city in all_cities:
            amt = sum(parts_map.get(p, {}).get(city, {}).get("amount", 0) for p in parts_map)
            vol = sum(parts_map.get(p, {}).get(city, {}).get("volume", 0) for p in parts_map)
            unit = amt / vol if vol > 0 else None
            if unit is not None:
                total_city[city] = {"amount": amt, "volume": vol, "unit": unit}
        if total_city:
            parts_map["总计"] = total_city
    if "总计" in parts_map and "主产品" in parts_map:
        sub_city = {}
        for city in all_cities:
            total_unit = parts_map.get("总计", {}).get(city, {}).get("unit")
            main_unit = parts_map.get("主产品", {}).get(city, {}).get("unit")
            if total_unit is None or main_unit is None:
                continue
            sub_city[city] = {"amount": 0.0, "volume": 0.0, "unit": total_unit - main_unit}
        if sub_city:
            parts_map["副产品"] = sub_city

    value_map = {}
    for part, cities in parts_map.items():
        value_map[part] = {c: data.get("unit") for c, data in cities.items()}
    return value_map, _order_cities(list(all_cities))


def _merge_city_order(base_rows: List[Dict[str, str]], *city_lists: List[str]) -> List[str]:
    seen = set()
    ordered = []
    for r in base_rows:
        city = r.get("城市") or r.get("工厂")
        if city and city not in seen:
            seen.add(city)
            ordered.append(city)
    for cities in city_lists:
        for city in cities:
            if city and city not in seen:
                seen.add(city)
                ordered.append(city)
    return _order_cities(ordered)


def _order_cities(cities: List[str]) -> List[str]:
    seen = set()
    ordered = []
    for city in CITY_ORDER:
        if city in cities and city not in seen:
            seen.add(city)
            ordered.append(city)
    for city in cities:
        if city and city not in seen:
            seen.add(city)
            ordered.append(city)
    return ordered


def _city_sort_index(city: str) -> int:
    try:
        return CITY_ORDER.index(city)
    except ValueError:
        return len(CITY_ORDER)


def _standard_rates_for_part(part: str) -> Tuple[float | None, float | None]:
    return PART_STANDARD_RATES.get(part, (None, None))


def _style_parts_pivot(pivot):
    def _highlight(data):
        styles = pd.DataFrame("", index=data.index, columns=data.columns)
        if STANDARD_RATE_COLUMNS[0] not in data.columns or STANDARD_RATE_COLUMNS[1] not in data.columns:
            return styles
        auto_col = STANDARD_RATE_COLUMNS[0]
        nonauto_col = STANDARD_RATE_COLUMNS[1]
        for city in AUTO_LINE_CITIES:
            if city in data.columns:
                mask = data[city].notna() & data[auto_col].notna() & (data[city] < data[auto_col])
                styles.loc[mask, city] = "color: red;"
        for city in NONAUTO_LINE_CITIES:
            if city in data.columns:
                mask = data[city].notna() & data[nonauto_col].notna() & (data[city] < data[nonauto_col])
                styles.loc[mask, city] = "color: red;"
        return styles

    return pivot.style.apply(_highlight, axis=None).format(
        lambda x: f"{x*100:.2f}%" if pd.notna(x) else ""
    )


def _style_value_pivot(pivot):
    target_cities = [c for c in CITY_ORDER if c in pivot.columns]

    def _highlight(data):
        styles = pd.DataFrame("", index=data.index, columns=data.columns)
        if not target_cities:
            return styles
        for idx, row in data.iterrows():
            series = row[target_cities].dropna()
            if series.empty:
                continue
            min_val = series.min()
            max_val = series.max()
            if min_val == max_val:
                continue
            for city in target_cities:
                val = row[city]
                if pd.isna(val):
                    continue
                if val == max_val:
                    styles.at[idx, city] = "color: green;"
                if val == min_val:
                    styles.at[idx, city] = "color: red;"
        return styles

    return pivot.style.apply(_highlight, axis=None).format(
        lambda x: _format_decimal2(x) if pd.notna(x) else ""
    )


def _sanitize_sheet_title(title: str, fallback: str) -> str:
    cleaned = str(title or "").strip()
    for ch in ["\\", "/", "?", "*", "[", "]", ":"]:
        cleaned = cleaned.replace(ch, "_")
    if not cleaned:
        cleaned = fallback
    return cleaned[:31]


def _build_full_report_bytes(
    core_records: List[Dict[str, str]],
    periods: List[str],
    part_records: List[Dict[str, str]],
    value_records: List[Dict[str, str]],
) -> bytes:
    from openpyxl import Workbook

    part_map, part_cities = _build_part_pivot_map(part_records)
    value_map, value_cities = _build_value_pivot_map(value_records, part_records)

    wb = Workbook()
    wb.remove(wb.active)
    used_titles = set()
    for idx, period in enumerate(periods):
        period_rows = [r for r in core_records if r.get("期间") == period]
        if not period_rows:
            continue
        base_title = _sanitize_sheet_title(period, f"Sheet{idx + 1}")
        title = base_title
        suffix = 2
        while title in used_titles:
            suffix_text = f"_{suffix}"
            title = f"{base_title[: 31 - len(suffix_text)]}{suffix_text}"
            suffix += 1
        used_titles.add(title)
        ws = wb.create_sheet(title=title)
        _write_report_sheet(ws, period_rows, part_map, value_map, part_cities, value_cities)

    if not wb.sheetnames:
        wb.create_sheet(title="Sheet1")
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _write_report_sheet(
    ws,
    core_rows: List[Dict[str, str]],
    part_map: Dict[str, Dict[str, float]],
    value_map: Dict[str, Dict[str, float]],
    part_cities: List[str],
    value_cities: List[str],
) -> None:
    from copy import copy
    from openpyxl.styles.colors import Color

    percent_fmt = "0.00%"
    value_fmt = "0.00"

    core_rows = sorted(
        core_rows,
        key=lambda r: (_city_sort_index(r.get("城市", "")), r.get("城市", "")),
    )

    row = 1
    ws.cell(row=row, column=1, value="月累计汇总")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    row += 1
    for col_idx, (header, _, _) in enumerate(SUMMARY_EXPORT_FIELDS, start=1):
        ws.cell(row=row, column=col_idx, value=header)
    data_start = row + 1
    for record in core_rows:
        row += 1
        for col_idx, (_, key, parser) in enumerate(SUMMARY_EXPORT_FIELDS, start=1):
            cell = ws.cell(row=row, column=col_idx, value=_coerce_value(record.get(key, ""), parser))
            if col_idx == 3 and cell.value != "":
                cell.number_format = percent_fmt
            elif col_idx == 5 and cell.value != "":
                cell.number_format = "#,##0"
            elif col_idx == 6 and cell.value != "":
                cell.number_format = "#,##0"
            elif col_idx == 7 and cell.value != "":
                cell.number_format = "0%"
            elif col_idx == 8 and cell.value != "":
                cell.number_format = "#,##0"
    data_end = row

    if data_end >= data_start:
        ws.merge_cells(start_row=data_start, start_column=2, end_row=data_end, end_column=2)
        for r in range(data_start + 1, data_end + 1):
            ws.cell(row=r, column=2, value=None)

    row += 1
    city_order = _merge_city_order(core_rows, part_cities, value_cities)
    part_city_start = 5
    part_city_end = part_city_start + len(city_order) - 1 if city_order else part_city_start - 1
    value_label_col = max(10, part_city_end + 2)
    value_city_start = value_label_col + 1
    value_city_end = value_city_start + len(city_order) - 1 if city_order else value_city_start - 1

    ws.cell(row=row, column=2, value="月产成细项对比")
    ws.cell(row=row, column=value_label_col, value="月产值细项对比")
    part_title_end = max(4, part_city_end)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=part_title_end)
    ws.merge_cells(start_row=row, start_column=value_label_col, end_row=row, end_column=max(value_label_col, value_city_end))

    row += 1
    ws.cell(row=row, column=2, value="部位")
    ws.cell(row=row, column=3, value="自动掏膛线工厂标准")
    ws.cell(row=row, column=4, value="非自动掏膛线工厂标准")
    red_color = Color(rgb="FFFF0000")
    green_color = Color(rgb="FF008000")
    value_targets = [c for c in CITY_ORDER if c in city_order]
    for idx, city in enumerate(city_order):
        ws.cell(row=row, column=part_city_start + idx, value=city)
    ws.cell(row=row, column=value_label_col, value="部位")
    for idx, city in enumerate(city_order):
        ws.cell(row=row, column=value_city_start + idx, value=city)

    col_widths = {
        "B": 15,
        "C": 12.6363636363636,
        "D": 14.0909090909091,
        "E": 13.6363636363636,
        "F": 10,
        "G": 10.2727272727273,
        "H": 13.1818181818182,
        "I": 12.9090909090909,
        "J": 17.2727272727273,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    for part in PART_ORDER:
        row += 1
        ws.cell(row=row, column=2, value=part)
        auto_rate, nonauto_rate = _standard_rates_for_part(part)
        cell = ws.cell(row=row, column=3, value=auto_rate if auto_rate is not None else "")
        if auto_rate is not None:
            cell.number_format = percent_fmt
        cell = ws.cell(row=row, column=4, value=nonauto_rate if nonauto_rate is not None else "")
        if nonauto_rate is not None:
            cell.number_format = percent_fmt
        value_candidates = [value_map.get(part, {}).get(c) for c in value_targets]
        value_candidates = [v for v in value_candidates if v is not None]
        value_min = min(value_candidates) if value_candidates else None
        value_max = max(value_candidates) if value_candidates else None
        value_span = value_min is not None and value_max is not None and value_min != value_max
        for idx, city in enumerate(city_order):
            val = part_map.get(part, {}).get(city)
            cell = ws.cell(row=row, column=part_city_start + idx, value=val if val is not None else "")
            if val is not None:
                cell.number_format = percent_fmt
            if val is not None and city in AUTO_LINE_CITIES and auto_rate is not None and val < auto_rate:
                font = copy(cell.font)
                font.color = red_color
                cell.font = font
            elif val is not None and city in NONAUTO_LINE_CITIES and nonauto_rate is not None and val < nonauto_rate:
                font = copy(cell.font)
                font.color = red_color
                cell.font = font
        value_label = part
        if part == "主产品":
            value_label = "主产品(整鸡还原)"
        elif part == "副产品":
            value_label = "副产品(整鸡还原)"
        ws.cell(row=row, column=value_label_col, value=value_label)
        for idx, city in enumerate(city_order):
            val = value_map.get(part, {}).get(city)
            cell = ws.cell(row=row, column=value_city_start + idx, value=val if val is not None else "")
            if val is not None:
                cell.number_format = value_fmt
            if value_span and city in value_targets and val is not None:
                if val == value_max:
                    font = copy(cell.font)
                    font.color = green_color
                    cell.font = font
                if val == value_min:
                    font = copy(cell.font)
                    font.color = red_color
                    cell.font = font

    row += 1
    ws.cell(row=row, column=8, value=REPORT_NOTE_RIGHT)
    for note in REPORT_NOTES_LEFT:
        row += 1
        ws.cell(row=row, column=2, value=note)


def main() -> None:
    st.set_page_config(page_title="城市 KPI 汇总", layout="wide")
    st.title("城市 KPI 汇总（多文件上传）")
    st.caption("上传多个 Excel（每个文件首个工作表，包含“日期”表头），自动汇总核心指标。")

    if CHART_PATH.exists():
        st.subheader("图表预览")
        st.image(str(CHART_PATH), use_column_width=True)

    uploaded_files = st.file_uploader("选择多个 .xlsx 文件", type=["xlsx"], accept_multiple_files=True)
    if not uploaded_files:
        st.info("等待上传文件。")
        return

    import re
    month_total_re = re.compile(r"^\s*(\d{1,2})月累计\s*$")

    def _is_month_total(text: str) -> bool:
        return bool(month_total_re.match(str(text or "")))

    records: List[Dict[str, str]] = []
    part_records: List[Dict[str, str]] = []
    value_records: List[Dict[str, str]] = []
    for f in uploaded_files:
        try:
            data = f.getvalue()
            recs = extract_records_from_bytes(data, f.name)
            records.extend(recs)
            part_records.extend(extract_monthly_parts_from_bytes(data, f.name))
            value_records.extend(extract_monthly_value_from_bytes(data, f.name))
        except Exception as exc:  # pragma: no cover - surfaced to user
            st.error(f"{f.name} 解析失败: {exc}")

    core_records_all = [r for r in records if _is_month_total(r.get("期间", ""))]
    if not records:
        st.warning("未解析到任何数据，请检查表头是否包含“日期”。")
        return

    if pd:
        df = pd.DataFrame(records, columns=SUMMARY_COLUMNS)
        core_df = df[df["期间"].apply(_is_month_total)].copy()
        if core_df.empty:
            st.info("未找到“X月累计”行。")
            return

        periods = sorted(core_df["期间"].dropna().unique().tolist())
        sel = st.multiselect("选择期间（可多选）", options=periods, default=periods)
        selected_periods = sel or periods
        core_view = core_df if not sel else core_df[core_df["期间"].isin(sel)]
        if core_view.empty:
            st.warning("当前选择下无数据。")
            return

        display_df = core_view.copy()
        display_df["_city_order"] = display_df["城市"].apply(_city_sort_index)
        display_df = display_df.sort_values(["期间", "_city_order", "城市"]).drop(columns=["_city_order"])
        for col in ["生肉产量(吨)", "销量(吨)", "宰鸡量(千只)"]:
            if col in display_df:
                display_df[col] = display_df[col].apply(_format_thousands0)
        if "产成率(%)" in display_df:
            display_df["产成率(%)"] = display_df["产成率(%)"].apply(_format_percent2)
        if "产销率(%)" in display_df:
            display_df["产销率(%)"] = display_df["产销率(%)"].apply(_format_percent0)

        st.subheader("月累计汇总（核心指标概览）")
        st.dataframe(display_df, use_container_width=True)
        csv_bytes = display_df.to_csv(index=False).encode("utf-8")
    else:
        core_records = [r for r in records if _is_month_total(r.get("期间", ""))]
        if not core_records:
            st.info("未找到“X月累计”行。")
            return
        periods = sorted({r.get("期间", "") for r in core_records if r.get("期间", "")})
        sel = st.multiselect("选择期间（可多选）", options=periods, default=periods)
        selected_periods = sel or periods
        core_view = [r for r in core_records if (not sel) or (r.get("期间", "") in sel)]
        if not core_view:
            st.warning("当前选择下无数据。")
            return
        core_view_sorted = sorted(
            core_view,
            key=lambda r: (r.get("期间", ""), _city_sort_index(r.get("城市", "")), r.get("城市", "")),
        )
        formatted_records = []
        for r in core_view_sorted:
            item = dict(r)
            item["生肉产量(吨)"] = _format_thousands0(r.get("生肉产量(吨)", ""))
            item["销量(吨)"] = _format_thousands0(r.get("销量(吨)", ""))
            item["宰鸡量(千只)"] = _format_thousands0(r.get("宰鸡量(千只)", ""))
            item["产成率(%)"] = _format_percent2(r.get("产成率(%)", ""))
            item["产销率(%)"] = _format_percent0(r.get("产销率(%)", ""))
            formatted_records.append(item)
        st.subheader("月累计汇总（核心指标概览）")
        st.dataframe(formatted_records, use_container_width=True, height=300)
        csv_bytes = to_csv_bytes(formatted_records)

    st.download_button(
        "下载汇总 CSV",
        data=csv_bytes,
        file_name="city_kpi_summary.csv",
        mime="text/csv",
        use_container_width=True,
    )

    # === 月产成细项对比 ===
    st.divider()
    st.subheader("月产成细项对比（本月至今累计）")
    if not part_records:
        st.info("未找到“本月至今累计”明细。")
        return
    # 归一化部位名称
    for r in part_records:
        r["部位"] = normalize_part_name(r.get("部位", ""))

    def _pivot_parts(records: List[Dict[str, str]]):
        if pd:
            dfp = pd.DataFrame(records)
            # 尝试解析小数
            dfp["产成率"] = dfp["产成率"].apply(_parse_percent)
            dfp = dfp.dropna(subset=["产成率"])
            dfp["部位"] = dfp["部位"].apply(normalize_part_name)
            dfp = dfp.groupby(["部位", "城市"], as_index=False)["产成率"].sum()
            pivot = dfp.pivot_table(index="部位", columns="城市", values="产成率", aggfunc="sum")
            # 补充派生行
            for new_part, parts in DERIVED_PARTS.items():
                parts_norm = _normalize_parts_list(parts)
                vals = {}
                for city in pivot.columns:
                    vals[city] = sum(
                        pivot.at[p, city] if p in pivot.index and pd.notna(pivot.at[p, city]) else 0 for p in parts_norm
                    )
                pivot.loc[new_part] = vals
            # 若已有“总计”则保留，否则求和
            if "总计" not in pivot.index:
                pivot.loc["总计"] = pivot.sum(min_count=1)
            if "总计" in pivot.index and "主产品" in pivot.index:
                pivot.loc["副产品"] = pivot.loc["总计"] - pivot.loc["主产品"]
            # 重新排序
            pivot = pivot.reindex(PART_ORDER, fill_value=pd.NA)
            pivot = pivot.reindex(columns=_order_cities(list(pivot.columns)))
            standards_df = pd.DataFrame(
                {
                    STANDARD_RATE_COLUMNS[0]: [
                        _standard_rates_for_part(part)[0] for part in PART_ORDER
                    ],
                    STANDARD_RATE_COLUMNS[1]: [
                        _standard_rates_for_part(part)[1] for part in PART_ORDER
                    ],
                },
                index=PART_ORDER,
            )
            pivot = pd.concat([standards_df, pivot], axis=1)
            return pivot
        else:
            # 手工 pivot
            parts = {}
            for r in records:
                part = r.get("部位")
                city = r.get("城市")
                val = _parse_percent(r.get("产成率"))
                if part and city and val is not None:
                    part_norm = normalize_part_name(part)
                    parts.setdefault(part_norm, {}).setdefault(city, 0.0)
                    parts[part_norm][city] += val

            # 派生行
            all_cities = {c for v in parts.values() for c in v}
            for new_part, comp in DERIVED_PARTS.items():
                comp_norm = _normalize_parts_list(comp)
                city_vals = {}
                for city in all_cities:
                    city_vals[city] = sum(parts.get(p, {}).get(city, 0) for p in comp_norm)
                parts[new_part] = city_vals

            # 总计
            if "总计" not in parts:
                parts["总计"] = {c: sum(parts.get(p, {}).get(c, 0) for p in parts if p != "总计") for c in all_cities}
            if "总计" in parts and "主产品" in parts:
                sub_city = {}
                for city in all_cities:
                    total_val = parts.get("总计", {}).get(city)
                    main_val = parts.get("主产品", {}).get(city)
                    if total_val is None or main_val is None:
                        continue
                    sub_city[city] = total_val - main_val
                if sub_city:
                    parts["副产品"] = sub_city

            rows = []
            all_cities = _order_cities(list({c for v in parts.values() for c in v}))
            for part in PART_ORDER:
                auto_rate, nonauto_rate = _standard_rates_for_part(part)
                row = {
                    "部位": part,
                    STANDARD_RATE_COLUMNS[0]: _format_percent2(auto_rate),
                    STANDARD_RATE_COLUMNS[1]: _format_percent2(nonauto_rate),
                }
                for c in all_cities:
                    val = parts.get(part, {}).get(c)
                    row[c] = f"{val*100:.2f}%" if val is not None else ""
                rows.append(row)
            return rows

    pivoted = _pivot_parts(part_records)
    if pd and isinstance(pivoted, pd.DataFrame):
        st.dataframe(_style_parts_pivot(pivoted), use_container_width=True)
        part_csv = (
            pivoted.applymap(lambda x: f"{x*100:.2f}%" if pd.notna(x) else "").to_csv().encode("utf-8")
        )
    else:
        st.dataframe(pivoted, use_container_width=True)
        part_csv = to_csv_bytes_generic(pivoted if isinstance(pivoted, list) else [])

    st.download_button(
        "下载月产成细项对比 CSV",
        data=part_csv,
        file_name="city_kpi_parts.csv",
        mime="text/csv",
        use_container_width=True,
    )

    # === 月产值细项对比（含税单价） ===
    st.divider()
    st.subheader("月产值细项对比（本月至今累计，含税单价）")
    if not value_records:
        st.info("未找到“本月至今累计”明细。")
        return

    rate_lookup = _build_rate_lookup(part_records)

    def _pivot_value(records: List[Dict[str, str]]):
        if pd:
            dfv = pd.DataFrame(records)
            dfv["含税单价"] = dfv.apply(
                lambda r: r["含税单价"]
                if pd.notna(r.get("含税单价"))
                else (r["含税金额"] / r["产量"] if pd.notna(r.get("含税金额")) and pd.notna(r.get("产量")) and r["产量"] > 0 else pd.NA),
                axis=1,
            )
            dfv = dfv.dropna(subset=["含税单价"])
            dfv["部位"] = dfv["部位"].apply(normalize_part_name)
            base = dfv[["部位", "城市", "含税单价", "产量", "含税金额"]]
            def _agg_base(group):
                volume = group["产量"].sum(min_count=1)
                amount = group["含税金额"].sum(min_count=1)
                weighted_amount = (group["含税单价"] * group["产量"]).sum(min_count=1)
                unit = pd.NA
                if pd.notna(volume) and volume > 0:
                    if pd.notna(amount):
                        unit = amount / volume
                    elif pd.notna(weighted_amount):
                        unit = weighted_amount / volume
                elif group["含税单价"].notna().any():
                    unit = group["含税单价"].mean()
                return pd.Series({"含税单价": unit, "产量": volume, "含税金额": amount})

            base = base.groupby(["部位", "城市"]).apply(_agg_base).reset_index()

            derived_rows = []
            for new_part, parts in DERIVED_PARTS.items():
                parts_norm = _normalize_parts_list(parts)
                if new_part == "主产品" and rate_lookup:
                    for city, part_rates in rate_lookup.items():
                        numerator = 0.0
                        denominator = 0.0
                        rate_sum = 0.0
                        total_rate = part_rates.get("总计")
                        for part in parts_norm:
                            rate = part_rates.get(part)
                            if rate is None:
                                continue
                            unit_series = base.loc[(base["部位"] == part) & (base["城市"] == city), "含税单价"]
                            if unit_series.empty:
                                continue
                            unit_val = unit_series.iloc[0]
                            if pd.isna(unit_val):
                                continue
                            numerator += float(unit_val) * rate
                            rate_sum += rate
                        denominator = total_rate if total_rate not in (None, 0) else rate_sum
                        if denominator > 0:
                            derived_rows.append(
                                {
                                    "部位": new_part,
                                    "城市": city,
                                    "含税单价": numerator / denominator,
                                    "产量": pd.NA,
                                    "含税金额": pd.NA,
                                }
                            )
                    continue
                sub = base[base["部位"].isin(parts_norm)]
                if sub.empty:
                    continue
                # Prefer weighted average from unit price + volume to avoid amount unit drift.
                unit_sub = sub.dropna(subset=["含税单价", "产量"]).copy()
                if not unit_sub.empty:
                    unit_sub["加权金额"] = unit_sub["含税单价"] * unit_sub["产量"]
                    agg = unit_sub.groupby("城市").agg({"加权金额": "sum", "产量": "sum"})
                    agg["含税单价"] = agg.apply(
                        lambda r: r["加权金额"] / r["产量"] if pd.notna(r["产量"]) and r["产量"] > 0 else pd.NA,
                        axis=1,
                    )
                    for city, row in agg.iterrows():
                        derived_rows.append(
                            {
                                "部位": new_part,
                                "城市": city,
                                "含税单价": row["含税单价"],
                                "产量": row["产量"],
                                "含税金额": row["加权金额"],
                            }
                        )
                    continue
                # Fallback to amount/volume if unit price is unavailable.
                amt_sub = sub.dropna(subset=["含税金额", "产量"])
                if amt_sub.empty:
                    continue
                agg = amt_sub.groupby("城市").agg({"含税金额": "sum", "产量": "sum"})
                agg["含税单价"] = agg.apply(
                    lambda r: r["含税金额"] / r["产量"] if pd.notna(r["产量"]) and r["产量"] > 0 else pd.NA,
                    axis=1,
                )
                for city, row in agg.iterrows():
                    derived_rows.append(
                        {
                            "部位": new_part,
                            "城市": city,
                            "含税单价": row["含税单价"],
                            "产量": row["产量"],
                            "含税金额": row["含税金额"],
                        }
                    )

            if derived_rows:
                derived_df = pd.DataFrame(derived_rows)
                base = base.merge(derived_df[["部位", "城市"]], on=["部位", "城市"], how="left", indicator=True)
                base = base[base["_merge"] == "left_only"].drop(columns=["_merge"])
                combined = pd.concat([base, derived_df], ignore_index=True)
            else:
                combined = base

            # 总计（仅 base 计算）
            if "总计" not in combined["部位"].values:
                total = base.groupby("城市").agg({"含税金额": "sum", "产量": "sum"})
                total["含税单价"] = total.apply(
                    lambda r: r["含税金额"] / r["产量"] if pd.notna(r["产量"]) and r["产量"] > 0 else pd.NA,
                    axis=1,
                )
                for city, row in total.iterrows():
                    combined = pd.concat(
                        [
                            combined,
                            pd.DataFrame(
                                [
                                    {
                                        "部位": "总计",
                                        "城市": city,
                                        "含税单价": row["含税单价"],
                                        "产量": row["产量"],
                                        "含税金额": row["含税金额"],
                                    }
                                ]
                            ),
                        ],
                        ignore_index=True,
                    )

            pivot = combined.pivot_table(index="部位", columns="城市", values="含税单价", aggfunc="first")
            if "总计" in pivot.index and "主产品" in pivot.index:
                pivot.loc["副产品"] = pivot.loc["总计"] - pivot.loc["主产品"]
            pivot = pivot.reindex(PART_ORDER, fill_value=pd.NA)
            pivot = pivot.reindex(columns=_order_cities(list(pivot.columns)))
            return pivot
        else:
            # 手工
            parts_map: Dict[str, Dict[str, Dict[str, float]]] = {}
            for r in records:
                part = normalize_part_name(r.get("部位"))
                city = r.get("城市")
                if not part or not city:
                    continue
                volume = _parse_number(r.get("产量"))
                amount = _parse_number(r.get("含税金额"))
                unit = _parse_number(r.get("含税单价"))
                if unit is None and amount is not None and volume is not None and volume > 0:
                    unit = amount / volume
                if unit is None:
                    continue
                parts_map.setdefault(part, {}).setdefault(city, {"amount": 0.0, "volume": 0.0})
                if amount is not None:
                    parts_map[part][city]["amount"] += amount
                if volume is not None:
                    parts_map[part][city]["volume"] += volume
                # store unit as fallback if no amount/volume
                parts_map[part][city]["unit"] = unit

            # 派生行
            all_cities = {c for v in parts_map.values() for c in v}
            for new_part, comp in DERIVED_PARTS.items():
                comp_norm = _normalize_parts_list(comp)
                if new_part == "主产品" and rate_lookup:
                    city_vals = {}
                    for city in all_cities:
                        numerator = 0.0
                        denominator = 0.0
                        rate_sum = 0.0
                        total_rate = rate_lookup.get(city, {}).get("总计")
                        amt = 0.0
                        vol = 0.0
                        for part in comp_norm:
                            part_data = parts_map.get(part, {}).get(city, {})
                            unit = part_data.get("unit")
                            rate = rate_lookup.get(city, {}).get(part)
                            if rate is not None and unit is not None:
                                numerator += unit * rate
                                rate_sum += rate
                            amt += part_data.get("amount", 0)
                            vol += part_data.get("volume", 0)
                        denominator = total_rate if total_rate not in (None, 0) else rate_sum
                        if denominator > 0:
                            city_vals[city] = {"amount": amt, "volume": vol, "unit": numerator / denominator}
                    if city_vals:
                        parts_map[new_part] = city_vals
                    continue
                city_vals = {}
                for city in all_cities:
                    amt = sum(parts_map.get(p, {}).get(city, {}).get("amount", 0) for p in comp_norm)
                    vol = sum(parts_map.get(p, {}).get(city, {}).get("volume", 0) for p in comp_norm)
                    unit = amt / vol if vol > 0 else None
                    if unit is not None:
                        city_vals[city] = {"amount": amt, "volume": vol, "unit": unit}
                if city_vals:
                    parts_map[new_part] = city_vals

            # 总计
            if "总计" not in parts_map:
                total_city = {}
                for city in all_cities:
                    amt = sum(parts_map.get(p, {}).get(city, {}).get("amount", 0) for p in parts_map)
                    vol = sum(parts_map.get(p, {}).get(city, {}).get("volume", 0) for p in parts_map)
                    unit = amt / vol if vol > 0 else None
                    if unit is not None:
                        total_city[city] = {"amount": amt, "volume": vol, "unit": unit}
                if total_city:
                    parts_map["总计"] = total_city
            if "总计" in parts_map and "主产品" in parts_map:
                sub_city = {}
                for city in all_cities:
                    total_unit = parts_map.get("总计", {}).get(city, {}).get("unit")
                    main_unit = parts_map.get("主产品", {}).get(city, {}).get("unit")
                    if total_unit is None or main_unit is None:
                        continue
                    sub_city[city] = {"amount": 0.0, "volume": 0.0, "unit": total_unit - main_unit}
                if sub_city:
                    parts_map["副产品"] = sub_city

            rows = []
            all_cities = _order_cities(list({c for v in parts_map.values() for c in v}))
            for part in PART_ORDER:
                row = {"部位": part}
                for c in all_cities:
                    unit = parts_map.get(part, {}).get(c, {}).get("unit")
                    row[c] = _format_decimal2(unit)
                rows.append(row)
            return rows

    pivot_v = _pivot_value(value_records)
    if pd and isinstance(pivot_v, pd.DataFrame):
        st.dataframe(_style_value_pivot(pivot_v), use_container_width=True)
        value_csv = (
            pivot_v.applymap(lambda x: _format_decimal2(x) if pd.notna(x) else "").to_csv().encode("utf-8")
        )
    else:
        st.dataframe(pivot_v, use_container_width=True)
        value_csv = to_csv_bytes_generic(pivot_v if isinstance(pivot_v, list) else [])

    st.download_button(
        "下载月产值细项对比 CSV",
        data=value_csv,
        file_name="city_kpi_value_parts.csv",
        mime="text/csv",
        use_container_width=True,
    )

    st.divider()
    st.subheader("下载全部报表（与产成及产值明细一致）")
    full_report_bytes = _build_full_report_bytes(
        core_records_all,
        selected_periods,
        part_records,
        value_records,
    )
    st.download_button(
        "下载全部报表 Excel",
        data=full_report_bytes,
        file_name="city_kpi_full_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


if __name__ == "__main__":
    main()
